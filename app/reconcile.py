"""Сверка: сколько писем реально доехало.

Соблазн большой — взять собственные счётчики переноса и объявить их отчётом.
Но они говорят лишь о том, сколько писем мы отправили в приёмник, а не о том,
сколько там лежит. Разница между этими двумя числами и есть весь смысл сверки:
письмо могло быть отвергнуто, съедено политикой хранения или не пройти по
размеру. Поэтому здесь мы честно опрашиваем приёмник и сравниваем с
инвентаризацией источника, снятой до переноса.

Отчёт выгружается в xlsx — это то, что показывают заказчику как доказательство,
что переезд состоялся.
"""

from __future__ import annotations

import logging
import threading
from concurrent.futures import ThreadPoolExecutor
from dataclasses import dataclass, field
from datetime import datetime, timezone
from io import BytesIO

from app.crypto import decrypt
from app.db import session_scope
from app.imap_probe import EndpointConfig, probe
from app.journal import log_event
from app.models import AUTH_MASTER, Endpoint, Mailbox, MailboxFolder, Project

log = logging.getLogger(__name__)


@dataclass
class ReconcileProgress:
    project_id: int
    total: int = 0
    checked: int = 0
    failed: int = 0
    running: bool = True
    current: str = ""
    finished_at: datetime | None = None

    @property
    def percent(self) -> int:
        if not self.total:
            return 0
        return int(round(100 * self.checked / self.total))


class ReconcileRunner:
    def __init__(self, project_id: int) -> None:
        self.project_id = project_id
        self.progress = ReconcileProgress(project_id=project_id)
        self._stop = threading.Event()
        self._lock = threading.Lock()
        self._thread: threading.Thread | None = None

    def start(self) -> None:
        self._thread = threading.Thread(
            target=self._run, name=f"reconcile-{self.project_id}", daemon=True
        )
        self._thread.start()

    def stop(self) -> None:
        self._stop.set()

    @property
    def is_alive(self) -> bool:
        return bool(self._thread and self._thread.is_alive())

    def _run(self) -> None:
        try:
            plan = self._prepare()
        except Exception:  # noqa: BLE001
            log.exception("Не удалось подготовить сверку проекта %s", self.project_id)
            self._finish()
            return

        if plan is None:
            self._finish()
            return

        mailbox_ids, cfg, parallel = plan
        with ThreadPoolExecutor(max_workers=parallel, thread_name_prefix="reconcile") as pool:
            futures = [pool.submit(self._check_one, mid, cfg) for mid in mailbox_ids]
            for future in futures:
                try:
                    future.result()
                except Exception:  # noqa: BLE001
                    log.exception("Сверка ящика завершилась исключением")

        self._finish()

    def _prepare(self):
        with session_scope() as session:
            project = session.get(Project, self.project_id)
            if project is None:
                return None
            dst = session.get(Endpoint, project.dst_endpoint_id) if project.dst_endpoint_id else None
            if dst is None:
                log_event(session, project_id=self.project_id, level="error",
                          code="reconcile_no_endpoint",
                          message="Сверка невозможна: не выбран сервер-приёмник.")
                return None

            mailboxes = (
                session.query(Mailbox)
                .filter(Mailbox.project_id == self.project_id)
                .order_by(Mailbox.id)
                .all()
            )
            selected = [m.id for m in mailboxes if _has_dst_credentials(m, dst)]
            without_creds = len(mailboxes) - len(selected)

            parallel = max(1, min(project.max_parallel, dst.max_parallel))
            with self._lock:
                self.progress.total = len(selected)

            log_event(
                session, project_id=self.project_id, code="reconcile_started",
                message=(
                    f"Сверка запущена: опрашиваем приёмник по {len(selected)} ящикам"
                    + (f", без пароля пропущено {without_creds}" if without_creds else "")
                    + "."
                ),
            )

            cfg = EndpointConfig(
                host=dst.host, port=dst.port, security=dst.security,
                verify_cert=dst.verify_cert, auth_mode=dst.auth_mode,
                master_username=dst.master_username,
                master_secret=decrypt(dst.master_secret_enc),
                master_separator=dst.master_separator,
            )

        return selected, cfg, parallel

    def _check_one(self, mailbox_id: int, cfg: EndpointConfig) -> None:
        if self._stop.is_set():
            return

        with session_scope() as session:
            mailbox = session.get(Mailbox, mailbox_id)
            if mailbox is None:
                return
            login = mailbox.dst_email
            secret = decrypt(mailbox.dst_password_enc)
            email = mailbox.src_email

        result = probe(cfg, login, secret, inventory=True, measure_size=True)

        with session_scope() as session:
            mailbox = session.get(Mailbox, mailbox_id)
            if mailbox is None:
                return

            if not result.ok:
                with self._lock:
                    self.progress.failed += 1
                log_event(
                    session, project_id=self.project_id, mailbox_id=mailbox_id,
                    level="error", code="reconcile_failed",
                    message=f"{email}: сверка не удалась — "
                            f"{result.error.message.lower() if result.error else 'неизвестная ошибка'}",
                )
            else:
                session.query(MailboxFolder).filter(
                    MailboxFolder.mailbox_id == mailbox_id, MailboxFolder.side == "dst"
                ).delete(synchronize_session=False)

                for folder in result.folders:
                    session.add(
                        MailboxFolder(
                            mailbox_id=mailbox_id, side="dst",
                            name_raw=folder.name_raw, name_display=folder.name_display,
                            delimiter=folder.delimiter, special_use=folder.special_use,
                            selectable=folder.selectable, messages=folder.messages,
                            size_bytes=folder.size_bytes, uidvalidity=folder.uidvalidity,
                        )
                    )

                mailbox.dst_total_messages = result.total_messages
                mailbox.dst_total_bytes = result.total_bytes
                mailbox.reconciled_at = datetime.now(timezone.utc)

                gap = _gap(mailbox.total_messages, result.total_messages)
                level, note = ("info", "расхождений нет") if gap == 0 else (
                    "warning", f"расхождение: {gap:+d} писем"
                )
                log_event(
                    session, project_id=self.project_id, mailbox_id=mailbox_id,
                    level=level, code="reconcile_ok",
                    message=f"{email}: на приёмнике "
                            f"{result.total_messages if result.total_messages is not None else '?'} "
                            f"писем, {note}",
                )

        with self._lock:
            self.progress.checked += 1
            self.progress.current = email

    def _finish(self) -> None:
        with self._lock:
            self.progress.running = False
            self.progress.finished_at = datetime.now(timezone.utc)
            checked, failed = self.progress.checked, self.progress.failed

        with session_scope() as session:
            log_event(
                session, project_id=self.project_id, code="reconcile_finished",
                message=f"Сверка завершена: опрошено {checked} ящиков, "
                        f"недоступно {failed}. Отчёт можно выгрузить в xlsx.",
            )


_runners: dict[int, ReconcileRunner] = {}
_registry_lock = threading.Lock()


def start_reconcile(project_id: int) -> ReconcileRunner:
    with _registry_lock:
        existing = _runners.get(project_id)
        if existing and existing.is_alive:
            return existing
        runner = ReconcileRunner(project_id)
        _runners[project_id] = runner
    runner.start()
    return runner


def get_reconcile(project_id: int) -> ReconcileRunner | None:
    with _registry_lock:
        return _runners.get(project_id)


# --- отчёт -----------------------------------------------------------------


def _gap(source: int | None, destination: int | None) -> int:
    if source is None or destination is None:
        return 0
    return destination - source


def _has_dst_credentials(mailbox: Mailbox, dst: Endpoint) -> bool:
    return bool(mailbox.dst_password_enc) or dst.auth_mode == AUTH_MASTER


def build_report(project_id: int) -> tuple[BytesIO, str]:
    """Собрать отчёт о сверке в xlsx."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    header_font = Font(bold=True)
    warn_fill = PatternFill("solid", fgColor="FFF3CD")
    bad_fill = PatternFill("solid", fgColor="F8D7DA")

    with session_scope() as session:
        project = session.get(Project, project_id)
        src = session.get(Endpoint, project.src_endpoint_id) if project.src_endpoint_id else None
        dst = session.get(Endpoint, project.dst_endpoint_id) if project.dst_endpoint_id else None
        mailboxes = (
            session.query(Mailbox)
            .filter(Mailbox.project_id == project_id)
            .order_by(Mailbox.src_email)
            .all()
        )

        workbook = Workbook()

        summary = workbook.active
        summary.title = "Сводка"
        summary.append(["Проект", project.name])
        summary.append(["Источник", src.name if src else "—"])
        summary.append(["Приёмник", dst.name if dst else "—"])
        summary.append(["Отчёт сформирован", datetime.now().strftime("%d.%m.%Y %H:%M")])
        summary.append([])
        summary.append(["Ящиков в проекте", len(mailboxes)])
        summary.append(["Перенесено", sum(1 for m in mailboxes if m.status == "done")])
        summary.append(["С ошибками", sum(1 for m in mailboxes if m.status == "failed")])
        summary.append(["Сверено с приёмником", sum(1 for m in mailboxes if m.reconciled_at)])
        summary.append([])
        summary.append([
            "Числа источника сняты до переноса; числа приёмника получены опросом "
            "самого сервера, а не подсчётом того, что мы отправили."
        ])
        for cell in summary["A"]:
            cell.font = header_font
        summary.column_dimensions["A"].width = 28
        summary.column_dimensions["B"].width = 60

        sheet = workbook.create_sheet("Ящики")
        columns = [
            ("Откуда", 34), ("Куда", 34), ("Комментарий", 22), ("Статус", 16),
            ("Писем на источнике", 18), ("Писем на приёмнике", 18), ("Разница", 12),
            ("Объём источника, МБ", 20), ("Объём приёмника, МБ", 20),
            ("Ошибка", 34), ("Сверено", 18),
        ]
        sheet.append([title for title, _ in columns])
        for index, (_, width) in enumerate(columns, start=1):
            sheet.cell(row=1, column=index).font = header_font
            sheet.cell(row=1, column=index).alignment = Alignment(wrap_text=True, vertical="top")
            sheet.column_dimensions[sheet.cell(row=1, column=index).column_letter].width = width

        statuses = {
            "done": "перенесён", "failed": "ошибка", "running": "идёт",
            "queued": "в очереди", "check_ok": "доступ подтверждён",
            "check_failed": "нет доступа", "new": "не проверялся",
        }

        for mailbox in mailboxes:
            gap = _gap(mailbox.total_messages, mailbox.dst_total_messages)
            sheet.append([
                mailbox.src_email,
                mailbox.dst_email,
                mailbox.note or "",
                statuses.get(mailbox.status, mailbox.status),
                mailbox.total_messages,
                mailbox.dst_total_messages,
                gap if mailbox.reconciled_at else None,
                round((mailbox.total_bytes or 0) / 1048576, 1) if mailbox.total_bytes else None,
                round((mailbox.dst_total_bytes or 0) / 1048576, 1) if mailbox.dst_total_bytes else None,
                mailbox.last_error or "",
                mailbox.reconciled_at.strftime("%d.%m.%Y %H:%M") if mailbox.reconciled_at else "",
            ])
            row_index = sheet.max_row
            if mailbox.status == "failed":
                for column in range(1, len(columns) + 1):
                    sheet.cell(row=row_index, column=column).fill = bad_fill
            elif mailbox.reconciled_at and gap < 0:
                # На приёмнике меньше, чем было на источнике, — то самое, ради
                # чего сверка и делается.
                for column in range(1, len(columns) + 1):
                    sheet.cell(row=row_index, column=column).fill = warn_fill

        sheet.freeze_panes = "A2"

        gaps = workbook.create_sheet("Расхождения по папкам")
        gaps.append(["Ящик", "Папка источника", "Писем на источнике", "Писем на приёмнике"])
        for cell in gaps[1]:
            cell.font = header_font
        gaps.column_dimensions["A"].width = 34
        gaps.column_dimensions["B"].width = 34
        gaps.column_dimensions["C"].width = 20
        gaps.column_dimensions["D"].width = 20

        for mailbox in mailboxes:
            if not mailbox.reconciled_at:
                continue
            if _gap(mailbox.total_messages, mailbox.dst_total_messages) == 0:
                continue

            folders = (
                session.query(MailboxFolder)
                .filter(MailboxFolder.mailbox_id == mailbox.id)
                .all()
            )
            dst_by_name = {
                f.name_display: f.messages for f in folders if f.side == "dst"
            }
            for folder in folders:
                if folder.side != "src":
                    continue
                gaps.append([
                    mailbox.src_email,
                    folder.name_display,
                    folder.messages,
                    dst_by_name.get(folder.name_display),
                ])
        gaps.freeze_panes = "A2"

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        safe_name = "".join(
            char if char.isalnum() or char in " -_" else "_" for char in project.name
        ).strip() or "project"
        filename = f"Сверка — {safe_name}.xlsx"

    return buffer, filename
