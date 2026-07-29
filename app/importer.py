"""Импорт списка ящиков из CSV/xlsx.

CSV в админской реальности — это не RFC 4180, а «то, что выплюнул Excel»:
cp1251, разделитель `;`, BOM в начале, пароли со спецсимволами без кавычек
и невидимые пробелы по краям. Поэтому импорт трёхшаговый:

    загрузка с автоопределением -> маппинг колонок -> превью с проблемами

и только потом запись в БД. Ни одна строка не попадает в проект, пока человек
не увидел, что именно приехало.
"""

from __future__ import annotations

import csv
import io
import logging
import re
from dataclasses import dataclass, field
from pathlib import Path

log = logging.getLogger(__name__)

# Поля, которые мы умеем принимать. Всё остальное (хост, порт, папки) живёт
# в эндпоинте и настройках проекта — иначе список на 200 строк превращается
# в конфигурационный файл, который невозможно проверить глазами.
FIELDS = ("src_email", "dst_email", "src_password", "dst_password", "note")
REQUIRED_FIELDS = ("src_email", "dst_email")

# Подсказки для автоугадывания колонок. Регистр и пробелы не важны.
COLUMN_HINTS: dict[str, tuple[str, ...]] = {
    "src_email": (
        "src_email", "source", "src", "откуда", "источник", "старый", "old",
        "from", "email_from", "old_email", "почта", "адрес", "login", "логин",
    ),
    "src_password": (
        "src_password", "src_pass", "password_src", "пароль_откуда", "старый_пароль",
        "old_password", "pass_from", "пароль источника",
    ),
    "dst_email": (
        "dst_email", "destination", "dst", "куда", "приёмник", "приемник", "новый",
        "new", "to", "email_to", "new_email", "target",
    ),
    "dst_password": (
        "dst_password", "dst_pass", "password_dst", "пароль_куда", "новый_пароль",
        "new_password", "pass_to", "пароль приёмника", "пароль приемника",
    ),
    "note": ("note", "comment", "комментарий", "примечание", "фио", "имя", "отдел", "name"),
}

EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

CANDIDATE_ENCODINGS = ("utf-8-sig", "cp1251", "utf-8")
CANDIDATE_DELIMITERS = ";,\t|"


# --------------------------------------------------------------------------
# Шаг 1. Разбор файла
# --------------------------------------------------------------------------


@dataclass
class ParsedTable:
    headers: list[str]
    rows: list[list[str]]
    encoding: str | None = None
    delimiter: str | None = None
    has_header: bool = True
    notes: list[str] = field(default_factory=list)

    @property
    def column_count(self) -> int:
        return len(self.headers)


def parse_upload(path: Path, filename: str | None = None) -> ParsedTable:
    name = (filename or path.name).lower()
    if name.endswith((".xlsx", ".xlsm")):
        return _parse_xlsx(path)
    return _parse_csv(path)


def _parse_xlsx(path: Path) -> ParsedTable:
    """xlsx предпочтительнее CSV: в нём попросту нет проблемы кодировок."""
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        raw_rows: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            # Значения ячеек НЕ обрезаем: хвостовой пробел в пароле — это то,
            # что мы обязаны показать человеку, а не молча съесть.
            cells = ["" if v is None else str(v) for v in row]
            if any(c.strip() for c in cells):
                raw_rows.append(cells)
    finally:
        wb.close()

    if not raw_rows:
        return ParsedTable(headers=[], rows=[], notes=["Файл пуст."])

    return _split_header(raw_rows, ParsedTable(headers=[], rows=[], encoding="xlsx"))


def _parse_csv(path: Path) -> ParsedTable:
    data = path.read_bytes()
    text, encoding, note = _decode(data)

    result = ParsedTable(headers=[], rows=[], encoding=encoding)
    if note:
        result.notes.append(note)

    delimiter = _sniff_delimiter(text)
    result.delimiter = delimiter

    reader = csv.reader(io.StringIO(text, newline=""), delimiter=delimiter)
    # Ячейки оставляем как есть — см. комментарий в _parse_xlsx.
    raw_rows = [row for row in reader if any(c.strip() for c in row)]

    if not raw_rows:
        result.notes.append("Файл пуст.")
        return result

    return _split_header(raw_rows, result)


def _decode(data: bytes) -> tuple[str, str, str | None]:
    """utf-8 пробуем первой: cp1251 почти никогда не падает и «успешно»
    расшифрует UTF-8 в кракозябры, если дать ей шанс первой."""
    for encoding in CANDIDATE_ENCODINGS:
        try:
            return data.decode(encoding), encoding, None
        except UnicodeDecodeError:
            continue

    return (
        data.decode("utf-8", errors="replace"),
        "utf-8 (с потерями)",
        "Кодировку определить не удалось, часть символов может быть испорчена. "
        "Надёжнее пересохранить файл в xlsx.",
    )


def _sniff_delimiter(text: str) -> str:
    sample = text[:8192]
    try:
        return csv.Sniffer().sniff(sample, delimiters=CANDIDATE_DELIMITERS).delimiter
    except csv.Error:
        pass

    # Sniffer регулярно ошибается на файлах с одной колонкой или пустыми
    # значениями — считаем сами по первой непустой строке.
    first_line = next((ln for ln in sample.splitlines() if ln.strip()), "")
    counts = {d: first_line.count(d) for d in CANDIDATE_DELIMITERS}
    best = max(counts, key=lambda d: counts[d])
    return best if counts[best] > 0 else ","


def _split_header(raw_rows: list[list[str]], result: ParsedTable) -> ParsedTable:
    first = raw_rows[0]
    # Строка с адресом почты — это данные, а не заголовок.
    has_header = not any("@" in cell for cell in first)
    result.has_header = has_header

    if has_header:
        result.headers = [cell.lstrip("﻿").strip() for cell in first]
        body = raw_rows[1:]
    else:
        result.headers = [f"Колонка {i + 1}" for i in range(len(first))]
        body = raw_rows
        result.notes.append("Строка заголовков не найдена, первая строка считается данными.")

    width = len(result.headers)
    normalized: list[list[str]] = []
    ragged = 0
    for row in body:
        if len(row) != width:
            ragged += 1
        row = row[:width] + [""] * max(0, width - len(row))
        normalized.append(row)

    if ragged:
        # Почти всегда это неэкранированный разделитель внутри пароля.
        result.notes.append(
            f"У {ragged} строк число колонок не совпадает с заголовком. "
            f"Обычно это разделитель «{result.delimiter or ';'}» внутри пароля — "
            "проверь такие строки в превью особенно внимательно."
        )

    result.rows = normalized
    return result


# --------------------------------------------------------------------------
# Шаг 2. Маппинг колонок
# --------------------------------------------------------------------------


# Заголовки в жизни выглядят как «Пароль откуда», «Почта куда», «e-mail (новый)».
# Списком подстрок это не покрыть, поэтому считаем по двум признакам отдельно:
# о какой стороне речь и что за величина.
SIDE_TOKENS = {
    "src": ("откуда", "источник", "старый", "стар", "old", "src", "from", "source"),
    "dst": ("куда", "приёмник", "приемник", "новый", "нов", "new", "dst", "to", "target"),
}
KIND_TOKENS = {
    "email": ("почта", "адрес", "ящик", "email", "e-mail", "mail", "логин", "login"),
    "password": ("пароль", "password", "pass", "пасс"),
    "note": ("фио", "имя", "сотрудник", "отдел", "коммент", "примечание", "note", "comment"),
}


def _contains(header: str, tokens: tuple[str, ...]) -> bool:
    return any(token in header for token in tokens)


def _score(header: str, field_name: str) -> int:
    """Насколько заголовок похож на это поле. Отрицательное значение — не годится."""
    if not header:
        return -1

    if header in COLUMN_HINTS.get(field_name, ()):
        return 100

    if field_name == "note":
        return 3 if _contains(header, KIND_TOKENS["note"]) else -1

    side, kind = field_name.split("_", 1)
    score = 0

    if _contains(header, KIND_TOKENS[kind]):
        score += 3
    elif kind == "password":
        # «Пароль» в заголовке колонки с паролем есть почти всегда. Если слова
        # нет — это скорее адрес, чем пароль без подписи.
        return -1

    # Колонка с паролем не должна достаться полю адреса и наоборот.
    other_kind = "email" if kind == "password" else "password"
    if _contains(header, KIND_TOKENS[other_kind]):
        return -1

    if _contains(header, SIDE_TOKENS[side]):
        score += 3
    elif _contains(header, SIDE_TOKENS["dst" if side == "src" else "src"]):
        return -1

    return score if score else -1


def guess_mapping(headers: list[str]) -> dict[str, int | None]:
    """Предположить, какая колонка чему соответствует.

    Это именно предположение: человек подтверждает его глазами на следующем шаге.
    """
    mapping: dict[str, int | None] = {f: None for f in FIELDS}
    normalized = [h.strip().casefold() for h in headers]

    candidates = []
    for field_index, field_name in enumerate(FIELDS):
        for index, header in enumerate(normalized):
            score = _score(header, field_name)
            if score > 0:
                # При равенстве очков порядок стабильный: сначала поля из FIELDS,
                # потом колонки слева направо.
                candidates.append((-score, field_index, index, field_name))

    used_columns: set[int] = set()
    for _, _, index, field_name in sorted(candidates):
        if mapping[field_name] is not None or index in used_columns:
            continue
        mapping[field_name] = index
        used_columns.add(index)

    # Заголовков может не быть вовсе (тогда они «Колонка 1», «Колонка 2»…).
    # В этом случае берём самый частый порядок: откуда, пароль, куда, пароль.
    if mapping["src_email"] is None and mapping["dst_email"] is None:
        for position, field_name in enumerate(("src_email", "src_password",
                                               "dst_email", "dst_password", "note")):
            if position < len(headers):
                mapping[field_name] = position

    return mapping


# --------------------------------------------------------------------------
# Шаг 3. Превью с проблемами
# --------------------------------------------------------------------------

SEVERITY_ERROR = "error"      # строку импортировать нельзя
SEVERITY_WARNING = "warning"  # импортировать можно, но человек должен увидеть


@dataclass
class Issue:
    code: str
    message: str
    severity: str = SEVERITY_ERROR


@dataclass
class PreviewRow:
    number: int
    src_email: str = ""
    dst_email: str = ""
    src_password: str = ""
    dst_password: str = ""
    note: str = ""
    issues: list[Issue] = field(default_factory=list)

    @property
    def importable(self) -> bool:
        return not any(i.severity == SEVERITY_ERROR for i in self.issues)

    @property
    def has_password_whitespace(self) -> bool:
        return any(i.code == "password_whitespace" for i in self.issues)


@dataclass
class Preview:
    rows: list[PreviewRow]
    notes: list[str] = field(default_factory=list)

    @property
    def total(self) -> int:
        return len(self.rows)

    @property
    def importable_count(self) -> int:
        return sum(1 for r in self.rows if r.importable)

    @property
    def error_count(self) -> int:
        return self.total - self.importable_count

    @property
    def whitespace_count(self) -> int:
        return sum(1 for r in self.rows if r.has_password_whitespace)


def build_preview(
    table: ParsedTable,
    mapping: dict[str, int | None],
    *,
    trim_passwords: bool = False,
) -> Preview:
    preview = Preview(rows=[], notes=list(table.notes))

    missing = [f for f in REQUIRED_FIELDS if mapping.get(f) is None]
    if missing:
        names = {"src_email": "откуда", "dst_email": "куда"}
        preview.notes.append(
            "Не указаны обязательные колонки: " + ", ".join(names[m] for m in missing)
        )
        return preview

    seen_src: dict[str, int] = {}
    seen_dst: dict[str, int] = {}

    for number, raw in enumerate(table.rows, start=1):
        row = PreviewRow(number=number)

        for field_name in FIELDS:
            index = mapping.get(field_name)
            value = raw[index] if index is not None and index < len(raw) else ""
            setattr(row, field_name, value)

        # Адреса чистим всегда: пробел вокруг адреса никогда не значим.
        row.src_email = row.src_email.strip()
        row.dst_email = row.dst_email.strip()

        _check_password_whitespace(row, "src_password", "источника", trim_passwords)
        _check_password_whitespace(row, "dst_password", "приёмника", trim_passwords)

        _check_email(row, "src_email", "откуда")
        _check_email(row, "dst_email", "куда")

        if row.src_email and row.src_email.casefold() == row.dst_email.casefold():
            row.issues.append(
                Issue(
                    "same_address",
                    "Адрес источника и приёмника совпадают",
                    SEVERITY_WARNING,
                )
            )

        _check_duplicate(row, "src_email", seen_src, "Адрес источника уже встречался в строке")
        _check_duplicate(row, "dst_email", seen_dst, "Адрес приёмника уже встречался в строке")

        preview.rows.append(row)

    return preview


def _check_password_whitespace(
    row: PreviewRow, attr: str, side: str, trim: bool
) -> None:
    """Самая злая проблема списка: хвостовой пробел в пароле невидим глазом,
    логин падает с «неверный пароль», и человек ищет несуществующую проблему."""
    value = getattr(row, attr)
    if not value:
        return
    if value != value.strip():
        if trim:
            setattr(row, attr, value.strip())
        else:
            row.issues.append(
                Issue(
                    "password_whitespace",
                    f"В пароле {side} есть пробел в начале или в конце",
                    SEVERITY_WARNING,
                )
            )


def _check_email(row: PreviewRow, attr: str, label: str) -> None:
    value = getattr(row, attr)
    if not value:
        row.issues.append(Issue(f"empty_{attr}", f"Не заполнен адрес «{label}»"))
    elif not EMAIL_RE.match(value):
        row.issues.append(Issue(f"invalid_{attr}", f"Адрес «{label}» выглядит некорректно"))


def _check_duplicate(
    row: PreviewRow, attr: str, seen: dict[str, int], message: str
) -> None:
    value = getattr(row, attr).casefold()
    if not value:
        return
    if value in seen:
        severity = SEVERITY_ERROR if attr == "src_email" else SEVERITY_WARNING
        row.issues.append(Issue(f"duplicate_{attr}", f"{message} {seen[value]}", severity))
    else:
        seen[value] = row.number


# --------------------------------------------------------------------------
# Генерация списка по шаблону
# --------------------------------------------------------------------------


def generate_from_addresses(addresses: list[str], template: str) -> ParsedTable:
    """Построить пары «откуда -> куда» по шаблону вида {local}@newdomain.ru.

    При переезде домена список нужен только для исключений, а не для всех
    двухсот строк — их можно получить отсюда.

    Доступные подстановки: {local}, {domain}, {email}.
    """
    rows: list[list[str]] = []
    for raw in addresses:
        src = raw.strip()
        if not src:
            continue
        local, _, domain = src.partition("@")
        try:
            dst = template.format(local=local, domain=domain, email=src)
        except (KeyError, IndexError):
            # Кривой шаблон не должен ронять страницу — проблему покажет превью.
            dst = ""
        rows.append([src, dst])

    return ParsedTable(
        headers=["src_email", "dst_email"],
        rows=rows,
        encoding="generated",
        has_header=True,
    )
